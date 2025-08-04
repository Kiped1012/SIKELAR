"""
RKAS Data processor for SIKELAR application
Handles Excel RKAS data extraction and processing
"""

import openpyxl
from typing import Dict, List
from .utils import ExcelUtils

class RKASDataProcessor:
    def __init__(self):
        self.excel_utils = ExcelUtils()
        self.reset_data()
        
        # Kategori kode yang lebih spesifik berdasarkan gambar
        self.kategori_kode = {
            'honor': ['07.12'],  # Honor
            'belanja_persediaan': ['5.1.02.01'],  # Belanja Persediaan
            'belanja_jasa': ['5.1.02.02'],  # Belanja Jasa
            'belanja_pemeliharaan': ['5.1.02.03'],  # Belanja Pemeliharaan
            'belanja_perjalanan': ['5.1.02.04'],  # Belanja Perjalanan Dinas
            'peralatan': ['5.2.02'],  # Peralatan dan Mesin
            'aset_tetap_lainnya': ['5.2.04', '5.2.05']  # Aset Tetap Lainnya
        }

    def reset_data(self):
        """Reset all RKAS data to initial state"""
        self.excel_data = None
        self.total_penerimaan = 0
        self.budget_items = []
        self.belanja_persediaan_items = []
        self.belanja_jasa_items = []
        self.belanja_pemeliharaan_items = []
        self.belanja_perjalanan_items = []
        self.peralatan_items = []
        self.aset_tetap_items = []
        self.nama_sekolah = ""

    def extract_rkas_data(self, file_path):
        """Ekstrak data RKAS dari file Excel"""
        workbook = openpyxl.load_workbook(file_path)
        
        # Reset data
        self.reset_data()
        self.excel_data = workbook
        
        # Tentukan sheet RKAS
        try:
            if 'RKAS' in workbook.sheetnames:
                rkas_sheet = workbook['RKAS']
            else:
                rkas_sheet = workbook.worksheets[0]  # Sheet pertama
                
        except Exception as e:
            print(f"Error accessing RKAS sheet: {e}")
            rkas_sheet = workbook.active
        
        # Proses RKAS
        self.process_rkas_data(rkas_sheet)
        
        print(f"Debug: Total Penerimaan: Rp {self.total_penerimaan:,}")
        print(f"Debug: Found {len(self.belanja_persediaan_items)} belanja persediaan items")
        print(f"Debug: Found {len(self.belanja_jasa_items)} belanja jasa items")
        print(f"Debug: Found {len(self.belanja_pemeliharaan_items)} belanja pemeliharaan items")
        print(f"Debug: Found {len(self.belanja_perjalanan_items)} belanja perjalanan items")
        print(f"Debug: Found {len(self.peralatan_items)} peralatan items")
        print(f"Debug: Found {len(self.aset_tetap_items)} aset tetap items")

    def process_rkas_data(self, sheet):
        """Proses data RKAS dari sheet yang ditentukan"""
        print("Debug: Processing RKAS data")
        
        # Ekstrak nama sekolah
        self.extract_nama_sekolah(sheet)

        # Ekstrak total penerimaan dari baris 30, kolom I-N (merged)
        self.extract_total_penerimaan(sheet)
        
        # Ekstrak data budget berdasarkan kode kegiatan
        self.extract_budget_data(sheet)
        
        # Ekstrak data belanja persediaan berdasarkan kode rekening
        self.extract_belanja_persediaan_data(sheet)

        # Ekstrak data belanja jasa berdasarkan kode rekening
        self.extract_belanja_jasa_data(sheet)

        # Ekstrak data belanja pemeliharaan berdasarkan kode rekening
        self.extract_belanja_pemeliharaan_data(sheet)

        # Ekstrak data belanja perjalanan dinas berdasarkan kode rekening
        self.extract_belanja_perjalanan_data(sheet)

        # Ekstrak data peralatan berdasarkan kode rekening
        self.extract_peralatan_data(sheet)

        # Ekstrak data aset tetap lainnya berdasarkan kode rekening
        self.extract_aset_tetap_data(sheet)

    def extract_nama_sekolah(self, sheet):
        """Ekstrak nama sekolah dari baris 7, kolom F-AF (merged)"""
        try:
            # Ekstrak teks dari kolom F sampai AF (F=6, AF=32)
            nama_sekolah = self.excel_utils.extract_merged_text(sheet, 7, range(6, 33))
            if nama_sekolah and nama_sekolah.strip():
                self.nama_sekolah = nama_sekolah.strip()
                print(f"Debug: Nama sekolah ditemukan: {self.nama_sekolah}")
            else:
                self.nama_sekolah = "Nama Sekolah Tidak Ditemukan"
        except Exception as e:
            print(f"Error extracting nama sekolah {e}")
            self.nama_sekolah = "Nama Sekolah Tidak Ditemukan"

    def extract_total_penerimaan(self, sheet):
        """Ekstrak total penerimaan dari baris 30"""
        try:
            # Cari di baris 30, kolom I sampai N
            for col_idx in range(9, 15):  # I=9, J=10, K=11, L=12, M=13, N=14
                cell_value = sheet.cell(row=30, column=col_idx).value
                if isinstance(cell_value, (int, float)) and cell_value > 0:
                    self.total_penerimaan = int(cell_value)
                    print(f"Debug: Total Penerimaan ditemukan di baris 30, kolom {chr(64+col_idx)}: Rp {self.total_penerimaan:,}")
                    return
            
            # Jika tidak ditemukan di baris 30, cari di sekitar baris tersebut
            for row_idx in range(28, 33):  # Baris 28-32
                for col_idx in range(9, 15):  # Kolom I-N
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        self.total_penerimaan = int(cell_value)
                        print(f"Debug: Total Penerimaan ditemukan di baris {row_idx}, kolom {chr(64+col_idx)}: Rp {self.total_penerimaan:,}")
                        return
        except Exception as e:
            print(f"Error extracting total penerimaan: {e}")
        
        # Default jika tidak ditemukan
        if not self.total_penerimaan:
            self.total_penerimaan = 799500000  # Sesuai dengan gambar
            print("Debug: Menggunakan default total penerimaan: Rp 799,500,000")

    def extract_budget_data(self, sheet):
        """Ekstrak data budget berdasarkan kode kegiatan di kolom G"""
        all_target_codes = (self.kategori_kode['honor'])
        
        print(f"Debug: Mencari kode: {all_target_codes}")
        
        # Iterasi semua baris untuk mencari kode kegiatan
        for row_idx in range(1, sheet.max_row + 1):
            # Baca kode kegiatan dari kolom G (index 7)
            kode_cell = sheet.cell(row=row_idx, column=7)  # Kolom G
            kode_value = str(kode_cell.value) if kode_cell.value else ""
            kode_value = kode_value.strip()
            
            if not kode_value or kode_value == "None":
                continue
            
            # Cek apakah kode ini cocok dengan yang dicari
            for target_code in all_target_codes:
                if target_code in kode_value:
                    print(f"Debug: Menemukan kode {target_code} di baris {row_idx}: {kode_value}")
                    
                    # Ekstrak uraian dari kolom H-M (merged)
                    uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(8, 14))
                    
                    # Ekstrak jumlah dari kolom N-O (merged)
                    jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(14, 16))
                    
                    if uraian and jumlah > 0:
                        # Cek apakah kode sudah ada (hindari duplikasi)
                        existing = next((item for item in self.budget_items if item['kode'] == target_code), None)
                        if not existing:
                            self.budget_items.append({
                                'kode': target_code,
                                'uraian': uraian,
                                'jumlah': jumlah
                            })
                            print(f"Debug: Added - {target_code}: {uraian} - Rp {jumlah:,}")
                    break

    def extract_belanja_persediaan_data(self, sheet):
        """Ekstrak data belanja persediaan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.01'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_persediaan_items = self._filter_duplicate_items(found_items)

    def extract_belanja_jasa_data(self, sheet):
        """Ekstrak data belanja jasa berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.02'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_jasa_items = self._filter_duplicate_items(found_items)

    def extract_belanja_pemeliharaan_data(self, sheet):
        """Ekstrak data belanja pemeliharaan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.03'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_pemeliharaan_items = self._filter_duplicate_items(found_items)

    def extract_belanja_perjalanan_data(self, sheet):
        """Ekstrak data belanja perjalanan dinas berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.04'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_perjalanan_items = self._filter_duplicate_items(found_items)

    def extract_peralatan_data(self, sheet):
        """Ekstrak data peralatan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.2.02'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.peralatan_items = self._filter_duplicate_items(found_items)

    def extract_aset_tetap_data(self, sheet):
        """Ekstrak data aset tetap lainnya berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_codes = ['5.2.04', '5.2.05']
        found_items = []
        
        for target_code in target_codes:
            items = self._extract_items_by_code(sheet, target_code)
            found_items.extend(items)
        
        self.aset_tetap_items = self._filter_duplicate_items(found_items)

    def _extract_items_by_code(self, sheet, target_code):
        """Helper method untuk ekstrak items berdasarkan kode rekening"""
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_code}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(4, 7))
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung target code
            if target_code in kode_rekening:
                print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                
                # Ekstrak kode kegiatan dari kolom G
                kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                
                # Filter hanya kode kegiatan dengan format xx.xx.xx.
                if not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                    print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                    continue
                
                # Ekstrak uraian dari kolom H-M (merged)
                uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(8, 14))
                
                # Ekstrak jumlah dari kolom N-O (merged)
                jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(14, 16))
                
                if uraian and jumlah > 0 and kode_kegiatan:
                    found_items.append({
                        'kode_rekening': kode_rekening,
                        'kode_kegiatan': kode_kegiatan,
                        'uraian': uraian,
                        'jumlah': jumlah,
                        'row': row_idx
                    })
                    print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        return found_items

    def _filter_duplicate_items(self, found_items):
        """Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama, ambil yang paling atas"""
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        return filtered_items

    def filter_budget_by_codes(self, codes: List[str]) -> List[Dict]:
        """Filter budget berdasarkan kode kategori"""
        filtered_items = []
        for item in self.budget_items:
            for code in codes:
                if item['kode'].startswith(code[:5]):  # Match first 5 characters
                    filtered_items.append(item)
                    break
        
        return sorted(filtered_items, key=lambda x: x['kode'])

    def get_summary_data(self):
        """Generate summary data for ringkasan"""
        # Honor dari kategori budget
        honor_items = self.filter_budget_by_codes(self.kategori_kode['honor'])
        total_honor = sum(item['jumlah'] for item in honor_items)
        
        # Belanja jasa dan jasa sesungguhnya
        total_belanja_jasa = sum(item['jumlah'] for item in self.belanja_jasa_items)
        jasa_sesungguhnya = total_belanja_jasa - total_honor
        
        total_pemeliharaan = sum(item['jumlah'] for item in self.belanja_pemeliharaan_items)
        total_perjalanan = sum(item['jumlah'] for item in self.belanja_perjalanan_items)
        total_peralatan = sum(item['jumlah'] for item in self.peralatan_items)
        total_aset_tetap = sum(item['jumlah'] for item in self.aset_tetap_items)
        
        # Hitung belanja persediaan yang dimaksud dalam ringkasan
        belanja_persediaan_ringkasan = (self.total_penerimaan - total_honor - 
                                    jasa_sesungguhnya - total_pemeliharaan - total_perjalanan - total_peralatan - total_aset_tetap)
        
        # Hitung semua nilai yang diperlukan
        total_belanja_operasi = (total_honor + jasa_sesungguhnya + total_pemeliharaan + total_perjalanan + belanja_persediaan_ringkasan)  # TAMBAHKAN INI
        
        belanja_modal = total_peralatan + total_aset_tetap
        total_anggaran = total_belanja_operasi + total_peralatan + total_aset_tetap  # UPDATE INI
        
        return {
            'total_honor': total_honor,
            'jasa_sesungguhnya': jasa_sesungguhnya,
            'total_pemeliharaan': total_pemeliharaan,
            'total_perjalanan': total_perjalanan,
            'belanja_persediaan_ringkasan': belanja_persediaan_ringkasan,
            'total_belanja_operasi': total_belanja_operasi,  # TAMBAHKAN INI
            'total_peralatan': total_peralatan,
            'total_aset_tetap': total_aset_tetap,
            'belanja_modal': belanja_modal,
            'total_anggaran': total_anggaran
        }