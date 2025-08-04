"""
BKU Data processor for SIKELAR application
Handles Excel BKU data extraction and processing
"""

import openpyxl
from typing import Dict, List
from .utils import ExcelUtils

class BKUDataProcessor:
    def __init__(self):
        self.excel_utils = ExcelUtils()
        self.reset_data()

    def reset_data(self):
        """Reset all BKU data to initial state"""
        self.bku_data_available = False
        self.bku_belanja_persediaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_belanja_pemeliharaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_belanja_perjalanan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_peralatan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_aset_tetap_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_belanja_jasa_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }

    def extract_bku_data(self, file_path):
        """Ekstrak data BKU dari file Excel"""
        workbook = openpyxl.load_workbook(file_path)
        
        # Reset data
        self.reset_data()
        
        # Tentukan sheet BKU
        try:
            if 'BKU' in workbook.sheetnames:
                bku_sheet = workbook['BKU']
            elif len(workbook.worksheets) > 1:
                bku_sheet = workbook.worksheets[1]  # Sheet kedua
            else:
                bku_sheet = None  # Tidak ada sheet BKU
                
        except Exception as e:
            print(f"Error accessing BKU sheet: {e}")
            bku_sheet = None
        
        # Proses BKU jika ada
        if bku_sheet:
            self.process_bku_data(bku_sheet)
        else:
            print("Debug: BKU sheet not found")

    def process_bku_data(self, sheet):
        """Proses data BKU dari sheet yang ditentukan - IMPLEMENTASI LENGKAP"""
        print("Debug: Processing BKU data")
        
        # Validasi sheet
        if not sheet:
            print("Debug: BKU sheet not found")
            self.bku_data_available = False
            return
        
        print(f"Debug: BKU sheet found with {sheet.max_row} rows and {sheet.max_column} columns")
        
        # Ekstrak data BKU untuk semua triwulan
        self.extract_bku_belanja_persediaan_data(sheet)
        self.extract_bku_belanja_pemeliharaan_data(sheet)
        self.extract_bku_belanja_perjalanan_data(sheet)
        self.extract_bku_peralatan_data(sheet)
        self.extract_bku_aset_tetap_data(sheet)
        self.extract_bku_belanja_jasa_data(sheet)
        
        self.bku_data_available = True

    def extract_bku_belanja_persediaan_data(self, sheet):
        """Ekstrak data realisasi belanja persediaan dari BKU untuk semua triwulan"""
        target_code = '5.1.02.01'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_persediaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_persediaan_data[triwulan].append(item)

    def extract_bku_belanja_pemeliharaan_data(self, sheet):
        """Ekstrak data realisasi belanja pemeliharaan dari BKU untuk semua triwulan"""
        target_code = '5.1.02.03'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_pemeliharaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU pemeliharaan item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
    def _group_and_sum_bku_items(self, raw_items):
        """Group dan sum items BKU berdasarkan key yang sama"""
        from collections import defaultdict
        
        grouped = defaultdict(lambda: {
            'tanggal': None,
            'kode_rekening': '',
            'kode_kegiatan': '',
            'uraian': '',
            'jumlah': 0,
            'count': 0
        })
        
        # Group berdasarkan kombinasi tanggal, kode_kegiatan, kode_rekening, dan uraian
        for item in raw_items:
            key = f"{item['tanggal']}_{item['kode_kegiatan']}_{item['kode_rekening']}_{item['uraian']}"
            
            if grouped[key]['tanggal'] is None:
                grouped[key]['tanggal'] = item['tanggal']
                grouped[key]['kode_rekening'] = item['kode_rekening']
                grouped[key]['kode_kegiatan'] = item['kode_kegiatan']
                grouped[key]['uraian'] = item['uraian']
            
            grouped[key]['jumlah'] += item['jumlah']
            grouped[key]['count'] += 1
        
        # Convert ke list
        result = []
        for key, data in grouped.items():
            if data['jumlah'] > 0:
                result.append({
                    'tanggal': data['tanggal'],
                    'kode_rekening': data['kode_rekening'],
                    'kode_kegiatan': data['kode_kegiatan'],
                    'uraian': data['uraian'],
                    'jumlah': data['jumlah']
                })
        
        return result

    def _get_triwulan_from_date(self, tanggal):
        """Tentukan triwulan berdasarkan tanggal"""
        if not tanggal:
            return None
        
        try:
            if hasattr(tanggal, 'month'):
                month = tanggal.month
            else:
                # Jika tanggal adalah string, parse dulu
                parsed_date = self._parse_date_string(str(tanggal))
                if not parsed_date:
                    return None
                month = parsed_date.month
            
            if 1 <= month <= 3:
                return 'Triwulan 1'
            elif 4 <= month <= 6:
                return 'Triwulan 2'
            elif 7 <= month <= 9:
                return 'Triwulan 3'
            elif 10 <= month <= 12:
                return 'Triwulan 4'
            else:
                return None
        except:
            return None

    def _is_triwulan_complete(self, tanggal, sheet):
        """Cek apakah triwulan sudah complete (untuk filtering)"""
        # Implementasi sederhana - selalu return True
        # Bisa dikustomisasi sesuai kebutuhan business logic
        return True

    def _parse_date_string(self, date_str):
        """Parse string tanggal ke datetime object"""
        import datetime
        
        if not date_str or str(date_str).strip() == '' or str(date_str).strip().lower() == 'none':
            return None
        
        # Jika sudah berupa datetime object
        if isinstance(date_str, datetime.datetime):
            return date_str.date()
        elif isinstance(date_str, datetime.date):
            return date_str
        
        # Clean the string
        date_str = str(date_str).strip()
        
        # Format yang akan dicoba
        date_formats = [
            '%d-%m-%Y',    # 27-02-2025
            '%d/%m/%Y',    # 27/02/2025
            '%Y-%m-%d',    # 2025-02-27
            '%d.%m.%Y',    # 27.02.2025
            '%d %m %Y',    # 27 02 2025
            '%m/%d/%Y',    # 02/27/2025 (US format)
            '%Y/%m/%d',    # 2025/02/27
        ]
        
        for fmt in date_formats:
            try:
                return datetime.datetime.strptime(date_str, fmt).date()
            except:
                continue
        
        # Manual parsing sebagai fallback
        try:
            # Coba split dengan berbagai separator
            for sep in ['-', '/', '.', ' ']:
                if sep in date_str:
                    parts = date_str.split(sep)
                    if len(parts) == 3:
                        # Try DD-MM-YYYY format first
                        try:
                            day, month, year = map(int, parts)
                            if 1 <= day <= 31 and 1 <= month <= 12 and year > 1900:
                                return datetime.date(year, month, day)
                        except:
                            pass
                        
                        # Try YYYY-MM-DD format
                        try:
                            year, month, day = map(int, parts)
                            if 1 <= day <= 31 and 1 <= month <= 12 and year > 1900:
                                return datetime.date(year, month, day)
                        except:
                            pass
                    break
        except:
            pass
        
        print(f"Warning: Could not parse date: {date_str}")
        return None

    # Method-method getter yang juga perlu ditambahkan
    def get_bku_belanja_persediaan_by_triwulan(self, triwulan):
        """Get data realisasi belanja persediaan berdasarkan triwulan"""
        if not self.bku_data_available:
            return []
        
        return self.bku_belanja_persediaan_data.get(triwulan, [])

    def get_bku_belanja_pemeliharaan_by_triwulan(self, triwulan):
        """Get data realisasi belanja pemeliharaan berdasarkan triwulan"""
        if not self.bku_data_available:
            return []
        
        return self.bku_belanja_pemeliharaan_data.get(triwulan, [])

    def get_bku_belanja_perjalanan_by_triwulan(self, triwulan):
        """Get data realisasi belanja perjalanan berdasarkan triwulan"""
        if not self.bku_data_available:
            return []
        
        return self.bku_belanja_perjalanan_data.get(triwulan, [])

    def get_bku_peralatan_by_triwulan(self, triwulan):
        """Get data realisasi peralatan berdasarkan triwulan"""
        if not self.bku_data_available:
            return []
        
        return self.bku_peralatan_data.get(triwulan, [])

    def get_bku_aset_tetap_by_triwulan(self, triwulan):
        """Get data realisasi aset tetap berdasarkan triwulan"""
        if not self.bku_data_available:
            return []
        
        return self.bku_aset_tetap_data.get(triwulan, [])

    def get_bku_belanja_jasa_by_triwulan(self, triwulan):
        """Get data realisasi belanja jasa berdasarkan triwulan"""
        if not self.bku_data_available:
            return []
        
        return self.bku_belanja_jasa_data.get(triwulan, [])

    def get_bku_summary_data_by_triwulan(self, triwulan):
        """Generate summary data BKU untuk triwulan tertentu - FIXED VERSION"""
        if not self.bku_data_available:
            return {
                'total_belanja_operasi_bku': 0,
                'total_honor_bku': 0,
                'jasa_sesungguhnya_bku': 0,
                'total_pemeliharaan_bku': 0,
                'total_perjalanan_bku': 0,
                'total_persediaan_bku': 0,
                'belanja_modal_bku': 0,
                'total_peralatan_bku': 0,
                'total_aset_tetap_bku': 0,
                'total_realisasi': 0
            }
        
        persediaan_items = self.get_bku_belanja_persediaan_by_triwulan(triwulan)
        pemeliharaan_items = self.get_bku_belanja_pemeliharaan_by_triwulan(triwulan)
        perjalanan_items = self.get_bku_belanja_perjalanan_by_triwulan(triwulan)
        peralatan_items = self.get_bku_peralatan_by_triwulan(triwulan)
        aset_tetap_items = self.get_bku_aset_tetap_by_triwulan(triwulan)
        jasa_items = self.get_bku_belanja_jasa_by_triwulan(triwulan)
        
        total_persediaan = sum(item['jumlah'] for item in persediaan_items)
        total_pemeliharaan = sum(item['jumlah'] for item in pemeliharaan_items)
        total_perjalanan = sum(item['jumlah'] for item in perjalanan_items)
        total_peralatan = sum(item['jumlah'] for item in peralatan_items)
        total_aset_tetap = sum(item['jumlah'] for item in aset_tetap_items)
        total_jasa = sum(item['jumlah'] for item in jasa_items)
        
        # Hitung honor dari jasa (berdasarkan kode kegiatan 07.12)
        total_honor = 0
        for item in jasa_items:
            if item['kode_kegiatan'].startswith('07.12'):
                total_honor += item['jumlah']
        
        # Hitung jasa sesungguhnya
        jasa_sesungguhnya = total_jasa - total_honor

        # Hitung total belanja operasi
        total_belanja_operasi = total_persediaan + total_pemeliharaan + total_perjalanan + total_jasa
        
        # Hitung belanja modal
        belanja_modal = total_peralatan + total_aset_tetap
        
        # Hitung total realisasi
        total_realisasi = total_belanja_operasi + belanja_modal
        
        # RETURN DENGAN KEY YANG BENAR sesuai dengan yang diakses di _display_bku_summary_for_triwulan
        return {
            'total_belanja_operasi_bku': total_belanja_operasi,
            'total_honor_bku': total_honor,  # Belum diimplementasikan di BKU processor
            'jasa_sesungguhnya_bku': jasa_sesungguhnya,
            'total_pemeliharaan_bku': total_pemeliharaan,
            'total_perjalanan_bku': total_perjalanan,
            'total_persediaan_bku': total_persediaan,
            'belanja_modal_bku': belanja_modal,
            'total_peralatan_bku': total_peralatan,
            'total_aset_tetap_bku': total_aset_tetap,
            'total_realisasi': total_realisasi
        }

    def get_all_triwulan_summary(self):
        """Get ringkasan untuk semua triwulan sekaligus"""
        summary = {}
        
        for triwulan in ['Triwulan 1', 'Triwulan 2', 'Triwulan 3', 'Triwulan 4']:
            summary[triwulan] = self.get_bku_summary_data_by_triwulan(triwulan)
        
        return summary
            # Group and sum items by date, kode_kegiatan, kode_rekening
    
        """
    Implementasi lengkap method extract untuk semua kategori BKU
    Tambahkan ke dalam class BKUDataProcessor
    """

    def extract_bku_belanja_perjalanan_data(self, sheet):
        """Ekstrak data realisasi belanja perjalanan dari BKU untuk semua triwulan"""
        target_code = '5.1.02.04'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_perjalanan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU perjalanan item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_perjalanan_data[triwulan].append(item)

    def extract_bku_peralatan_data(self, sheet):
        """Ekstrak data realisasi peralatan dari BKU untuk semua triwulan"""
        target_code = '5.2.02'
        
        # Initialize storage untuk semua triwulan
        self.bku_peralatan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU peralatan item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_peralatan_data[triwulan].append(item)

    def extract_bku_aset_tetap_data(self, sheet):
        """Ekstrak data realisasi aset tetap lainnya dari BKU untuk semua triwulan"""
        target_codes = ['5.2.04', '5.2.05']
        
        # Initialize storage untuk semua triwulan
        self.bku_aset_tetap_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_codes}")
        
        # Collect all raw data first
        raw_items = []
        
        for target_code in target_codes:
            # Iterasi semua baris untuk mencari kode rekening
            for row_idx in range(1, sheet.max_row + 1):
                # Ekstrak kode rekening dari kolom F-G (merged)
                kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
                
                if not kode_rekening or target_code not in kode_rekening:
                    continue
                
                # Ekstrak tanggal dari kolom A-C (merged)
                tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
                if not tanggal_str:
                    continue
                
                # Parse tanggal
                try:
                    tanggal = self._parse_date_string(tanggal_str)
                    if not tanggal:
                        continue
                except:
                    continue
                
                # Ekstrak kode kegiatan dari kolom D-E (merged)
                kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
                if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                    continue
                
                # Ekstrak uraian dari kolom K-M (merged)
                uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
                if not uraian:
                    continue
                
                # Skip baris dengan kata "Terima" atau "Setor"
                if "terima" in uraian.lower() or "setor" in uraian.lower():
                    print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                    continue
                
                # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
                jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
                if jumlah <= 0:
                    continue
                
                raw_items.append({
                    'tanggal': tanggal,
                    'kode_rekening': kode_rekening,
                    'kode_kegiatan': kode_kegiatan,
                    'uraian': uraian,
                    'jumlah': jumlah,
                    'row': row_idx
                })
                
                print(f"Debug: Found BKU aset tetap item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_aset_tetap_data[triwulan].append(item)

    def extract_bku_belanja_jasa_data(self, sheet):
        """Ekstrak data realisasi belanja jasa dari BKU untuk semua triwulan"""
        target_code = '5.1.02.02'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_jasa_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU jasa item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_jasa_data[triwulan].append(item)

    def extract_bku_belanja_pemeliharaan_data(self, sheet):
        """Ekstrak data realisasi belanja pemeliharaan dari BKU untuk semua triwulan - FIXED"""
        target_code = '5.1.02.03'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_pemeliharaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU pemeliharaan item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # TAMBAHKAN INI: Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # TAMBAHKAN INI: Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_pemeliharaan_data[triwulan].append(item)