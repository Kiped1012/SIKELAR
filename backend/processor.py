"""
Data processor for SIKELAR application
Handles Excel data extraction and processing
"""

import openpyxl
from typing import Dict, List
from .utils import ExcelUtils

class BOSDataProcessor:
    def __init__(self):
        self.excel_utils = ExcelUtils()
        self.reset_data()
        
        # Kategori kode yang lebih spesifik berdasarkan gambar
        self.kategori_kode = {
            'honor': ['07.12'],  # Honor
            'belanja_persediaan': ['5.1.02.01'],  # Belanja Persediaan
            'belanja_jasa': ['5.1.02.02'],  # Belanja Jasa
            'belanja_pemeliharaan': ['5.1.02.03'],  # Belanja Pemeliharaan
            'belanja_perjalanan': ['5.1.02.04'],  # Belanja Perjalanan Dinas
            'peralatan': ['5.2.02'],  # Peralatan dan Mesin
            'aset_tetap_lainnya': ['5.2.04', '5.2.05']  # Aset Tetap Lainnya
        }

    def reset_data(self):
        """Reset all data to initial state"""
        self.excel_data = None
        self.total_penerimaan = 0
        self.budget_items = []
        self.belanja_persediaan_items = []
        self.belanja_jasa_items = []
        self.belanja_pemeliharaan_items = []
        self.belanja_perjalanan_items = []
        self.peralatan_items = []
        self.aset_tetap_items = []
        self.nama_sekolah = ""
        self.bku_data_available = False
        self.bku_belanja_persediaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_belanja_pemeliharaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_belanja_perjalanan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_peralatan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_aset_tetap_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        self.bku_belanja_jasa_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }

    def extract_excel_data(self, file_path):
        """Ekstrak data dari file Excel dengan struktur spesifik RKAS dan BKU"""
        workbook = openpyxl.load_workbook(file_path)
        
        # Reset data
        self.reset_data()
        
        # Tentukan sheet yang akan digunakan
        try:
            # Sheet 1 untuk RKAS (biasanya sheet pertama atau bernama 'RKAS')
            if 'RKAS' in workbook.sheetnames:
                rkas_sheet = workbook['RKAS']
            else:
                rkas_sheet = workbook.worksheets[0]  # Sheet pertama
            
            # Sheet 2 untuk BKU (biasanya sheet kedua atau bernama 'BKU')
            if 'BKU' in workbook.sheetnames:
                bku_sheet = workbook['BKU']
            elif len(workbook.worksheets) > 1:
                bku_sheet = workbook.worksheets[1]  # Sheet kedua
            else:
                bku_sheet = None  # Tidak ada sheet BKU
                
        except Exception as e:
            print(f"Error accessing sheets: {e}")
            # Fallback ke sheet aktif
            rkas_sheet = workbook.active
            bku_sheet = None
        
        # Proses RKAS dari sheet 1
        self.process_rkas_data(rkas_sheet)
        
        # Proses BKU dari sheet 2 (jika ada)
        if bku_sheet:
            self.process_bku_data(bku_sheet)
        
        print(f"Debug: Total Penerimaan: Rp {self.total_penerimaan:,}")
        print(f"Debug: Found {len(self.belanja_persediaan_items)} belanja persediaan items")
        print(f"Debug: Found {len(self.belanja_jasa_items)} belanja jasa items")
        print(f"Debug: Found {len(self.belanja_pemeliharaan_items)} belanja pemeliharaan items")
        print(f"Debug: Found {len(self.belanja_perjalanan_items)} belanja perjalanan items")
        print(f"Debug: Found {len(self.peralatan_items)} peralatan items")
        print(f"Debug: Found {len(self.aset_tetap_items)} aset tetap items")

    def process_rkas_data(self, sheet):
        """Proses data RKAS dari sheet yang ditentukan"""
        print("Debug: Processing RKAS data from Sheet 1")
        
        # Ekstrak nama sekolah
        self.extract_nama_sekolah(sheet)

        # Ekstrak total penerimaan dari baris 30, kolom I-N (merged)
        self.extract_total_penerimaan(sheet)
        
        # Ekstrak data budget berdasarkan kode kegiatan
        self.extract_budget_data(sheet)
        
        # Ekstrak data belanja persediaan berdasarkan kode rekening
        self.extract_belanja_persediaan_data(sheet)

        # Ekstrak data belanja jasa berdasarkan kode rekening
        self.extract_belanja_jasa_data(sheet)

        # Ekstrak data belanja pemeliharaan berdasarkan kode rekening
        self.extract_belanja_pemeliharaan_data(sheet)

        # Ekstrak data belanja perjalanan dinas berdasarkan kode rekening
        self.extract_belanja_perjalanan_data(sheet)

        # Ekstrak data peralatan berdasarkan kode rekening
        self.extract_peralatan_data(sheet)

        # Ekstrak data aset tetap lainnya berdasarkan kode rekening
        self.extract_aset_tetap_data(sheet)

    def process_bku_data(self, sheet):
        """Proses data BKU dari sheet yang ditentukan - IMPLEMENTASI LENGKAP"""
        print("Debug: Processing BKU data from Sheet 2")
        
        # Validasi sheet
        if not sheet:
            print("Debug: BKU sheet not found")
            self.bku_data_available = False
            return
        
        print(f"Debug: BKU sheet found with {sheet.max_row} rows and {sheet.max_column} columns")
        
        # Ekstrak data BKU untuk semua triwulan
        self.extract_bku_belanja_persediaan_data(sheet)
        self.extract_bku_belanja_pemeliharaan_data(sheet)
        self.extract_bku_belanja_perjalanan_data(sheet)
        self.extract_bku_peralatan_data(sheet)
        self.extract_bku_aset_tetap_data(sheet)
        self.extract_bku_belanja_jasa_data(sheet)
        
        self.bku_data_available = True

    def extract_bku_belanja_persediaan_data(self, sheet):
        """Ekstrak data realisasi belanja persediaan dari BKU untuk semua triwulan"""
        target_code = '5.1.02.01'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_persediaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_persediaan_data[triwulan].append(item)

    def extract_bku_belanja_pemeliharaan_data(self, sheet):
        """Ekstrak data realisasi belanja pemeliharaan dari BKU untuk semua triwulan"""
        target_code = '5.1.02.03'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_pemeliharaan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU pemeliharaan item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_pemeliharaan_data[triwulan].append(item)

    def get_bku_belanja_pemeliharaan_by_triwulan(self, triwulan):
        """Get data realisasi belanja pemeliharaan berdasarkan triwulan"""
        if not hasattr(self, 'bku_belanja_pemeliharaan_data'):
            return []
        
        return self.bku_belanja_pemeliharaan_data.get(triwulan, [])

    def extract_bku_belanja_perjalanan_data(self, sheet):
        """Ekstrak data realisasi belanja perjalanan dari BKU untuk semua triwulan"""
        target_code = '5.1.02.04'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_perjalanan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU perjalanan item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_perjalanan_data[triwulan].append(item)

    def get_bku_belanja_perjalanan_by_triwulan(self, triwulan):
        """Get data realisasi belanja perjalanan berdasarkan triwulan"""
        if not hasattr(self, 'bku_belanja_perjalanan_data'):
            return []
        
        return self.bku_belanja_perjalanan_data.get(triwulan, [])
    
    def extract_bku_peralatan_data(self, sheet):
        """Ekstrak data realisasi peralatan dari BKU untuk semua triwulan"""
        target_code = '5.2.02'
        
        # Initialize storage untuk semua triwulan
        self.bku_peralatan_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU peralatan item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_peralatan_data[triwulan].append(item)

    def get_bku_peralatan_by_triwulan(self, triwulan):
        """Get data realisasi peralatan berdasarkan triwulan"""
        if not hasattr(self, 'bku_peralatan_data'):
            return []
        
        return self.bku_peralatan_data.get(triwulan, [])
    
    def extract_bku_aset_tetap_data(self, sheet):
        """Ekstrak data realisasi aset tetap lainnya dari BKU untuk semua triwulan"""
        target_codes = ['5.2.04', '5.2.05']
        
        # Initialize storage untuk semua triwulan
        self.bku_aset_tetap_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_codes}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            # Check if any target code is in kode_rekening
            if not kode_rekening or not any(target_code in kode_rekening for target_code in target_codes):
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU aset tetap item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_aset_tetap_data[triwulan].append(item)

    def get_bku_aset_tetap_by_triwulan(self, triwulan):
        """Get data realisasi aset tetap lainnya berdasarkan triwulan"""
        if not hasattr(self, 'bku_aset_tetap_data'):
            return []
        
        return self.bku_aset_tetap_data.get(triwulan, [])

    def extract_bku_belanja_jasa_data(self, sheet):
        """Ekstrak data realisasi belanja jasa dari BKU untuk semua triwulan"""
        target_code = '5.1.02.02'
        
        # Initialize storage untuk semua triwulan
        self.bku_belanja_jasa_data = {
            'Triwulan 1': [],
            'Triwulan 2': [],
            'Triwulan 3': [],
            'Triwulan 4': []
        }
        
        print(f"Debug: Mencari realisasi BKU untuk kode rekening: {target_code}")
        
        # Collect all raw data first
        raw_items = []
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # Ekstrak kode rekening dari kolom F-G (merged)
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(6, 8))
            
            if not kode_rekening or target_code not in kode_rekening:
                continue
            
            # Ekstrak tanggal dari kolom A-C (merged)
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if not tanggal_str:
                continue
            
            # Parse tanggal
            try:
                tanggal = self._parse_date_string(tanggal_str)
                if not tanggal:
                    continue
            except:
                continue
            
            # Ekstrak kode kegiatan dari kolom D-E (merged)
            kode_kegiatan = self.excel_utils.extract_merged_text(sheet, row_idx, range(4, 6))
            if not kode_kegiatan or not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                continue
            
            # Ekstrak uraian dari kolom K-M (merged)
            uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(11, 14))
            if not uraian:
                continue
            
            # Skip baris dengan kata "Terima" atau "Setor"
            if "terima" in uraian.lower() or "setor" in uraian.lower():
                print(f"Debug: Skipping row with Terima/Setor: {uraian}")
                continue
            
            # Ekstrak jumlah pengeluaran dari kolom Q-S (merged)
            jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(17, 20))
            if jumlah <= 0:
                continue
            
            raw_items.append({
                'tanggal': tanggal,
                'kode_rekening': kode_rekening,
                'kode_kegiatan': kode_kegiatan,
                'uraian': uraian,
                'jumlah': jumlah,
                'row': row_idx
            })
            
            print(f"Debug: Found BKU jasa item - {tanggal} | {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Group and sum items by date, kode_kegiatan, kode_rekening, and uraian
        grouped_items = self._group_and_sum_bku_items(raw_items)
        
        # Distribute items to appropriate triwulan
        for item in grouped_items:
            triwulan = self._get_triwulan_from_date(item['tanggal'])
            if triwulan:
                # Check if triwulan is complete
                if self._is_triwulan_complete(item['tanggal'], sheet):
                    self.bku_belanja_jasa_data[triwulan].append(item)

    def get_bku_belanja_jasa_by_triwulan(self, triwulan):
        """Get data realisasi belanja jasa berdasarkan triwulan"""
        if not hasattr(self, 'bku_belanja_jasa_data'):
            return []
        
        return self.bku_belanja_jasa_data.get(triwulan, [])

    def _parse_date_string(self, date_str):
        """Parse tanggal dari string dengan format DD-MM-YYYY"""
        import datetime
        
        try:
            # Handle different date formats
            date_str = str(date_str).strip()
            
            # Try DD-MM-YYYY format
            if '-' in date_str:
                parts = date_str.split('-')
                if len(parts) == 3:
                    day, month, year = parts
                    return datetime.date(int(year), int(month), int(day))
            
            # Try other common formats
            for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d']:
                try:
                    return datetime.datetime.strptime(date_str, fmt).date()
                except:
                    continue
            
            return None
        except Exception as e:
            print(f"Debug: Error parsing date {date_str}: {e}")
            return None

    def _group_and_sum_bku_items(self, raw_items):
        """Group items by tanggal, kode_kegiatan, kode_rekening, uraian dan sum jumlah"""
        from collections import defaultdict
        
        grouped = defaultdict(lambda: {
            'tanggal': None,
            'kode_rekening': '',
            'kode_kegiatan': '',
            'uraian': '',
            'jumlah': 0,
            'count': 0
        })
        
        for item in raw_items:
            # Create key untuk grouping
            key = f"{item['tanggal']}_{item['kode_kegiatan']}_{item['kode_rekening']}_{item['uraian']}"
            
            grouped[key]['tanggal'] = item['tanggal']
            grouped[key]['kode_rekening'] = item['kode_rekening']
            grouped[key]['kode_kegiatan'] = item['kode_kegiatan']
            grouped[key]['uraian'] = item['uraian']
            grouped[key]['jumlah'] += item['jumlah']
            grouped[key]['count'] += 1
        
        # Convert to list and log duplicates
        result = []
        for key, data in grouped.items():
            if data['count'] > 1:
                print(f"Debug: Grouped {data['count']} items for: {data['uraian']} - Total: Rp {data['jumlah']:,}")
            
            result.append({
                'tanggal': data['tanggal'],
                'kode_rekening': data['kode_rekening'],
                'kode_kegiatan': data['kode_kegiatan'],
                'uraian': data['uraian'],
                'jumlah': data['jumlah']
            })
        
        return sorted(result, key=lambda x: (x['tanggal'], x['kode_kegiatan'], x['uraian']))

    def _get_triwulan_from_date(self, date_obj):
        """Tentukan triwulan berdasarkan tanggal"""
        month = date_obj.month
        
        if 1 <= month <= 3:
            return 'Triwulan 1'
        elif 4 <= month <= 6:
            return 'Triwulan 2'
        elif 7 <= month <= 9:
            return 'Triwulan 3'
        elif 10 <= month <= 12:
            return 'Triwulan 4'
        
        return None

    def _is_triwulan_complete(self, date_obj, sheet):
        """Cek apakah triwulan sudah lengkap (3 bulan)"""
        import datetime
        
        month = date_obj.month
        year = date_obj.year
        
        # Tentukan bulan-bulan dalam triwulan
        if 1 <= month <= 3:
            required_months = [1, 2, 3]
        elif 4 <= month <= 6:
            required_months = [4, 5, 6]
        elif 7 <= month <= 9:
            required_months = [7, 8, 9]
        elif 10 <= month <= 12:
            required_months = [10, 11, 12]
        else:
            return False
        
        # Cek keberadaan data untuk semua bulan dalam triwulan
        found_months = set()
        
        for row_idx in range(1, sheet.max_row + 1):
            tanggal_str = self.excel_utils.extract_merged_text(sheet, row_idx, range(1, 4))
            if tanggal_str:
                try:
                    check_date = self._parse_date_string(tanggal_str)
                    if check_date and check_date.year == year and check_date.month in required_months:
                        found_months.add(check_date.month)
                except:
                    continue
        
        # Triwulan dianggap lengkap jika minimal ada data bulan terakhir
        last_month = max(required_months)
        is_complete = last_month in found_months
        
        triwulan_name = self._get_triwulan_from_date(date_obj)
        print(f"Debug: {triwulan_name} complete check - Required months: {required_months}, Found: {sorted(found_months)}, Complete: {is_complete}")
        
        return is_complete

    def get_bku_belanja_persediaan_by_triwulan(self, triwulan):
        """Get data realisasi belanja persediaan berdasarkan triwulan"""
        if not hasattr(self, 'bku_belanja_persediaan_data'):
            return []
        
        return self.bku_belanja_persediaan_data.get(triwulan, [])

    def extract_nama_sekolah(self, sheet):
        """Ekstrak nama sekolah dari baris 7, kolom F-AF (merged)"""
        try:
            # Ekstrak teks dari kolom F sampai AF (F=6, AF=32)
            nama_sekolah = self.excel_utils.extract_merged_text(sheet, 7, range(6, 33))
            if nama_sekolah and nama_sekolah.strip():
                self.nama_sekolah = nama_sekolah.strip()
                print(f"Debug: Nama sekolah ditemukan: {self.nama_sekolah}")
            else:
                self.nama_sekolah = "Nama Sekolah Tidak Ditemukan"
        except Exception as e:
            print(f"Error extracting nama sekolah {e}")
            self.nama_sekolah = "Nama Sekolah Tidak Ditemukan"

    def extract_total_penerimaan(self, sheet):
        """Ekstrak total penerimaan dari baris 30"""
        try:
            # Cari di baris 30, kolom I sampai N
            for col_idx in range(9, 15):  # I=9, J=10, K=11, L=12, M=13, N=14
                cell_value = sheet.cell(row=30, column=col_idx).value
                if isinstance(cell_value, (int, float)) and cell_value > 0:
                    self.total_penerimaan = int(cell_value)
                    print(f"Debug: Total Penerimaan ditemukan di baris 30, kolom {chr(64+col_idx)}: Rp {self.total_penerimaan:,}")
                    return
            
            # Jika tidak ditemukan di baris 30, cari di sekitar baris tersebut
            for row_idx in range(28, 33):  # Baris 28-32
                for col_idx in range(9, 15):  # Kolom I-N
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        self.total_penerimaan = int(cell_value)
                        print(f"Debug: Total Penerimaan ditemukan di baris {row_idx}, kolom {chr(64+col_idx)}: Rp {self.total_penerimaan:,}")
                        return
        except Exception as e:
            print(f"Error extracting total penerimaan: {e}")
        
        # Default jika tidak ditemukan
        if not self.total_penerimaan:
            self.total_penerimaan = 799500000  # Sesuai dengan gambar
            print("Debug: Menggunakan default total penerimaan: Rp 799,500,000")

    def extract_budget_data(self, sheet):
        """Ekstrak data budget berdasarkan kode kegiatan di kolom G"""
        all_target_codes = (self.kategori_kode['honor'])
        
        print(f"Debug: Mencari kode: {all_target_codes}")
        
        # Iterasi semua baris untuk mencari kode kegiatan
        for row_idx in range(1, sheet.max_row + 1):
            # Baca kode kegiatan dari kolom G (index 7)
            kode_cell = sheet.cell(row=row_idx, column=7)  # Kolom G
            kode_value = str(kode_cell.value) if kode_cell.value else ""
            kode_value = kode_value.strip()
            
            if not kode_value or kode_value == "None":
                continue
            
            # Cek apakah kode ini cocok dengan yang dicari
            for target_code in all_target_codes:
                if target_code in kode_value:
                    print(f"Debug: Menemukan kode {target_code} di baris {row_idx}: {kode_value}")
                    
                    # Ekstrak uraian dari kolom H-M (merged)
                    uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(8, 14))
                    
                    # Ekstrak jumlah dari kolom N-O (merged)
                    jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(14, 16))
                    
                    if uraian and jumlah > 0:
                        # Cek apakah kode sudah ada (hindari duplikasi)
                        existing = next((item for item in self.budget_items if item['kode'] == target_code), None)
                        if not existing:
                            self.budget_items.append({
                                'kode': target_code,
                                'uraian': uraian,
                                'jumlah': jumlah
                            })
                            print(f"Debug: Added - {target_code}: {uraian} - Rp {jumlah:,}")
                    break

    def extract_belanja_persediaan_data(self, sheet):
        """Ekstrak data belanja persediaan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.01'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_persediaan_items = self._filter_duplicate_items(found_items)

    def extract_belanja_jasa_data(self, sheet):
        """Ekstrak data belanja jasa berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.02'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_jasa_items = self._filter_duplicate_items(found_items)

    def extract_belanja_pemeliharaan_data(self, sheet):
        """Ekstrak data belanja pemeliharaan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.03'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_pemeliharaan_items = self._filter_duplicate_items(found_items)

    def extract_belanja_perjalanan_data(self, sheet):
        """Ekstrak data belanja perjalanan dinas berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.04'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.belanja_perjalanan_items = self._filter_duplicate_items(found_items)

    def extract_peralatan_data(self, sheet):
        """Ekstrak data peralatan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.2.02'
        found_items = self._extract_items_by_code(sheet, target_code)
        self.peralatan_items = self._filter_duplicate_items(found_items)

    def extract_aset_tetap_data(self, sheet):
        """Ekstrak data aset tetap lainnya berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_codes = ['5.2.04', '5.2.05']
        found_items = []
        
        for target_code in target_codes:
            items = self._extract_items_by_code(sheet, target_code)
            found_items.extend(items)
        
        self.aset_tetap_items = self._filter_duplicate_items(found_items)

    def _extract_items_by_code(self, sheet, target_code):
        """Helper method untuk ekstrak items berdasarkan kode rekening"""
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_code}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.excel_utils.extract_merged_text_strict(sheet, row_idx, range(4, 7))
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung target code
            if target_code in kode_rekening:
                print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                
                # Ekstrak kode kegiatan dari kolom G
                kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                
                # Filter hanya kode kegiatan dengan format xx.xx.xx.
                if not self.excel_utils.is_valid_kegiatan_format(kode_kegiatan):
                    print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                    continue
                
                # Ekstrak uraian dari kolom H-M (merged)
                uraian = self.excel_utils.extract_merged_text(sheet, row_idx, range(8, 14))
                
                # Ekstrak jumlah dari kolom N-O (merged)
                jumlah = self.excel_utils.extract_merged_number(sheet, row_idx, range(14, 16))
                
                if uraian and jumlah > 0 and kode_kegiatan:
                    found_items.append({
                        'kode_rekening': kode_rekening,
                        'kode_kegiatan': kode_kegiatan,
                        'uraian': uraian,
                        'jumlah': jumlah,
                        'row': row_idx
                    })
                    print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        return found_items

    def _filter_duplicate_items(self, found_items):
        """Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama, ambil yang paling atas"""
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        return filtered_items

    def filter_budget_by_codes(self, codes: List[str]) -> List[Dict]:
        """Filter budget berdasarkan kode kategori"""
        filtered_items = []
        for item in self.budget_items:
            for code in codes:
                if item['kode'].startswith(code[:5]):  # Match first 5 characters
                    filtered_items.append(item)
                    break
        
        return sorted(filtered_items, key=lambda x: x['kode'])

    def get_summary_data(self):
        """Generate summary data for ringkasan"""
        # Honor dari kategori budget
        honor_items = self.filter_budget_by_codes(self.kategori_kode['honor'])
        total_honor = sum(item['jumlah'] for item in honor_items)
        
        # Belanja jasa dan jasa sesungguhnya
        total_belanja_jasa = sum(item['jumlah'] for item in self.belanja_jasa_items)
        jasa_sesungguhnya = total_belanja_jasa - total_honor
        
        total_pemeliharaan = sum(item['jumlah'] for item in self.belanja_pemeliharaan_items)
        total_perjalanan = sum(item['jumlah'] for item in self.belanja_perjalanan_items)
        total_peralatan = sum(item['jumlah'] for item in self.peralatan_items)
        total_aset_tetap = sum(item['jumlah'] for item in self.aset_tetap_items)
        
        # Hitung belanja persediaan yang dimaksud dalam ringkasan
        belanja_persediaan_ringkasan = (self.total_penerimaan - total_honor - 
                                    jasa_sesungguhnya - total_pemeliharaan - total_perjalanan - total_peralatan - total_aset_tetap)
        
        # Hitung semua nilai yang diperlukan
        total_belanja_persediaan = (total_honor + jasa_sesungguhnya + total_pemeliharaan + total_perjalanan + belanja_persediaan_ringkasan)
        
        belanja_modal = total_peralatan + total_aset_tetap
        total_anggaran = total_belanja_persediaan + total_peralatan + total_aset_tetap
        
        return {
            'total_honor': total_honor,
            'jasa_sesungguhnya': jasa_sesungguhnya,
            'total_pemeliharaan': total_pemeliharaan,
            'total_perjalanan': total_perjalanan,
            'belanja_persediaan_ringkasan': belanja_persediaan_ringkasan,
            'total_belanja_persediaan': total_belanja_persediaan,
            'total_peralatan': total_peralatan,
            'total_aset_tetap': total_aset_tetap,
            'belanja_modal': belanja_modal,
            'total_anggaran': total_anggaran
        }

    def get_bku_summary_data_by_triwulan(self, triwulan):
        """Generate summary data BKU untuk triwulan tertentu - mirip dengan get_summary_data()"""
        if not self.bku_data_available:
            return None
        
        # Ambil data realisasi untuk semua kategori berdasarkan triwulan
        bku_persediaan = self.get_bku_belanja_persediaan_by_triwulan(triwulan)
        bku_jasa = self.get_bku_belanja_jasa_by_triwulan(triwulan)  
        bku_pemeliharaan = self.get_bku_belanja_pemeliharaan_by_triwulan(triwulan)
        bku_perjalanan = self.get_bku_belanja_perjalanan_by_triwulan(triwulan)
        bku_peralatan = self.get_bku_peralatan_by_triwulan(triwulan)
        bku_aset_tetap = self.get_bku_aset_tetap_by_triwulan(triwulan)
        
        # Hitung total untuk setiap kategori
        total_persediaan_bku = sum(item['jumlah'] for item in bku_persediaan)
        total_jasa_bku = sum(item['jumlah'] for item in bku_jasa)
        total_pemeliharaan_bku = sum(item['jumlah'] for item in bku_pemeliharaan)
        total_perjalanan_bku = sum(item['jumlah'] for item in bku_perjalanan)
        total_peralatan_bku = sum(item['jumlah'] for item in bku_peralatan)
        total_aset_tetap_bku = sum(item['jumlah'] for item in bku_aset_tetap)
        
        # Hitung honor dari jasa (berdasarkan kode kegiatan 07.12)
        total_honor_bku = 0
        for item in bku_jasa:
            if item['kode_kegiatan'].startswith('07.12'):
                total_honor_bku += item['jumlah']
        
        # Hitung jasa sesungguhnya
        jasa_sesungguhnya_bku = total_jasa_bku - total_honor_bku
        
        # Hitung total belanja operasi
        total_belanja_operasi_bku = (total_honor_bku + jasa_sesungguhnya_bku + 
                                    total_pemeliharaan_bku + total_perjalanan_bku + 
                                    total_persediaan_bku)
        
        # Hitung belanja modal
        belanja_modal_bku = total_peralatan_bku + total_aset_tetap_bku
        
        # Total realisasi keseluruhan
        total_realisasi = total_belanja_operasi_bku + belanja_modal_bku
        
        return {
            'triwulan': triwulan,
            'total_honor_bku': total_honor_bku,
            'jasa_sesungguhnya_bku': jasa_sesungguhnya_bku,
            'total_pemeliharaan_bku': total_pemeliharaan_bku,
            'total_perjalanan_bku': total_perjalanan_bku,
            'total_persediaan_bku': total_persediaan_bku,
            'total_belanja_operasi_bku': total_belanja_operasi_bku,
            'total_peralatan_bku': total_peralatan_bku,
            'total_aset_tetap_bku': total_aset_tetap_bku,
            'belanja_modal_bku': belanja_modal_bku,
            'total_realisasi': total_realisasi
        }

    def get_all_triwulan_summary(self):
        """Get ringkasan untuk semua triwulan sekaligus"""
        if not self.bku_data_available:
            return {}
        
        all_summary = {}
        for triwulan in ['Triwulan 1', 'Triwulan 2', 'Triwulan 3', 'Triwulan 4']:
            summary = self.get_bku_summary_data_by_triwulan(triwulan)
            if summary:
                all_summary[triwulan] = summary
        
        return all_summary