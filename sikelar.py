import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from typing import Dict, List
import re

class BOSBudgetAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("SIKELAR")
        self.root.geometry("1000x700")
        self.root.configure(bg='#f0f0f0')
        
        self.excel_data = None
        self.total_penerimaan = 0
        self.budget_items = []
        self.belanja_persediaan_items = []  # Items untuk belanja persediaan
        self.belanja_jasa_items = []  # Items untuk belanja jasa
        self.belanja_pemeliharaan_items = []  # Items untuk belanja pemeliharaan
        self.belanja_perjalanan_items = []  # Items untuk belanja perjalanan dinas
        self.peralatan_items = []  # Items untuk peralatan
        self.aset_tetap_items = []  # Items untuk aset tetap lainnya
        self.nama_sekolah = ""
        
        # Kategori kode yang lebih spesifik berdasarkan gambar
        self.kategori_kode = {
            'buku': ['05.02.01', '05.02.02', '05.02.03', '05.02.04', '05.02.05'],  # Pengadaan Buku
            'sarana_prasarana': ['05.08.01', '05.08.08', '05.08.12'],  # Sarana Prasarana
            'honor': ['07.12'],  # Honor
            'belanja_persediaan': ['5.1.02.01'],  # Belanja Persediaan
            'belanja_jasa': ['5.1.02.02'],  # Belanja Jasa
            'belanja_pemeliharaan': ['5.1.02.03'],  # Belanja Pemeliharaan
            'belanja_perjalanan': ['5.1.02.04'],  # Belanja Perjalanan Dinas
            'peralatan': ['5.2.02'],  # Peralatan dan Mesin
            'aset_tetap_lainnya': ['5.2.04', '5.2.05']  # Aset Tetap Lainnya
        }
        
        self.setup_ui()
    
    def setup_ui(self):
        header_frame = tk.Frame(self.root, bg='#2c3e50', height=80)
        header_frame.pack(fill='x', padx=10, pady=(10,0))
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text="Sistem Informasi Pengelompokan Anggaran dan Rekening", 
                            font=('Arial', 16, 'bold'), fg='white', bg='#2c3e50')
        title_label.pack(expand=True)
        
        upload_frame = tk.Frame(self.root, bg='#ecf0f1', relief='raised', bd=2)
        upload_frame.pack(fill='x', padx=10, pady=10)
        
        upload_label = tk.Label(upload_frame, text="Upload File Excel RKAS:", 
                            font=('Arial', 12, 'bold'), bg='#ecf0f1')
        upload_label.pack(pady=10)
        
        # Pindahkan upload_btn_frame ke dalam upload_frame
        upload_btn_frame = tk.Frame(upload_frame, bg="#ecf0f1")
        upload_btn_frame.pack(pady=10)

        # Gunakan font seragam untuk tombol
        common_font = ('Arial', 11, 'bold')

        # Tombol Upload
        self.upload_btn = tk.Button(upload_btn_frame, text="Pilih File Excel (.xlsx)", command=self.upload_excel,
                            bg='#3498db', fg='white', font=common_font,
                            padx=20, pady=5, cursor='hand2')
        self.upload_btn.pack(side='left', padx=(0, 5))

        # Tombol Reset, simbol ✖ besar
        reset_btn = tk.Button(upload_btn_frame, text="✖", command=self.reset_data,
                            bg='red', fg='white', font=common_font,
                            padx=5, pady=5, cursor='hand2', bd=1, relief='raised')
        reset_btn.pack(side='left')

        
        self.file_label = tk.Label(upload_frame, text="Belum ada file yang dipilih", 
                                font=('Arial', 10), bg='#ecf0f1', fg='#7f8c8d')
        self.file_label.pack(pady=5)
        
        # Membagi button frame menjadi dua baris
        button_frame1 = tk.Frame(self.root, bg='#f0f0f0')
        button_frame1.pack(fill='x', padx=10, pady=(10, 5))
        
        button_frame2 = tk.Frame(self.root, bg='#f0f0f0')
        button_frame2.pack(fill='x', padx=10, pady=(5, 10))
        
        # Baris pertama buttons
        tk.Button(button_frame1, text="Buku (05.02.)", command=self.show_buku,
                bg='#e74c3c', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=15).pack(side='left', padx=10)
        
        tk.Button(button_frame1, text="Sarana & Prasarana (05.08.)", command=self.show_sarana_prasarana,
                bg='#f39c12', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=20).pack(side='left', padx=10)
        
        tk.Button(button_frame1, text="Honor (07.12.)", command=self.show_honor,
                bg='#27ae60', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=15).pack(side='left', padx=10)
        
        # Button Belanja Persediaan di samping Honor pada baris yang sama
        tk.Button(button_frame1, text="Pakai Habis (5.1.02.01)", command=self.show_belanja_persediaan,
                bg='#9b59b6', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=25).pack(side='left', padx=10)

        # Button Belanja Jasa tepat di samping Belanja Persediaan
        tk.Button(button_frame1, text="Jasa (5.1.02.02)", command=self.show_belanja_jasa,
                bg='#e67e22', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=25).pack(side='left', padx=10)
        
        # Button Belanja Pemeliharaan
        tk.Button(button_frame1, text="Pemeliharaan (5.1.02.03)", command=self.show_belanja_pemeliharaan,
                bg="#050607", fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=25).pack(side='left', padx=10)
        
        # Button Belanja Perjalanan Dinas
        tk.Button(button_frame1, text="Perjalanan Dinas (5.1.02.04)", command=self.show_belanja_perjalanan,
                bg='#e67e22', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=25).pack(side='left', padx=10)
        
        # Button Peralatan
        tk.Button(button_frame2, text="Peralatan dan Mesin (5.2.02)", command=self.show_peralatan,
                bg='#e67e22', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=20).pack(side='left', padx=10)
        
        # Button Aset Tetap Lainnya
        tk.Button(button_frame2, text="Aset Tetap Lainnya (5.2.04 & 5.2.05)", command=self.show_aset_tetap,
                bg='#2980b9', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=30).pack(side='left', padx=10)
        
        tk.Button(button_frame2, text="Ringkasan", command=self.show_ringkasan,
                bg='#16a085', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10, cursor='hand2', width=15).pack(side='left', padx=10)

        
        self.results_frame = tk.Frame(self.root, bg='#ffffff', relief='sunken', bd=2)
        self.results_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.create_results_table()
    
    def create_results_table(self):
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        self.result_title = tk.Label(self.results_frame, text="Pilih kategori untuk melihat rincian anggaran", 
                                    font=('Arial', 16, 'bold'), bg='#ffffff')
        self.result_title.pack(pady=20)
        
        self.table_frame = tk.Frame(self.results_frame, bg='#ffffff')
        self.table_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Default columns untuk kategori biasa
        columns = ('Kode Kegiatan', 'Uraian', 'Jumlah (Rp)')
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show='headings', height=15)
        
        # Style untuk memperbesar font tabel
        style = ttk.Style()
        style.configure("Treeview", font=('Arial', 12))
        style.configure("Treeview.Heading", font=('Arial', 14, 'bold'))
        
        for col in columns:
            self.tree.heading(col, text=col)
        self.tree.column('Kode Kegiatan', width=120, anchor='center')
        self.tree.column('Uraian', width=400, anchor='w')
        self.tree.column('Jumlah (Rp)', width=150, anchor='w')  # Diubah dari 'e' ke 'w' untuk rata kiri
        
        scrollbar = ttk.Scrollbar(self.table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self.summary_frame = tk.Frame(self.results_frame, bg='#ecf0f1', relief='raised', bd=1)
        self.summary_frame.pack(fill='x', padx=20, pady=10)
        
        self.total_label = tk.Label(self.summary_frame, text="", font=('Arial', 14, 'bold'), bg='#ecf0f1')
        self.total_label.pack(pady=5)
        
        self.percentage_label = tk.Label(self.summary_frame, text="", font=('Arial', 14, 'bold'), bg='#ecf0f1')
        self.percentage_label.pack(pady=5)
    
    def create_belanja_persediaan_table(self):
        """Buat tabel khusus untuk belanja persediaan dengan kolom yang berbeda"""
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        self.result_title = tk.Label(self.results_frame, text="Rincian Pakai Habis (5.1.02.01)", 
                                    font=('Arial', 16, 'bold'), bg='#ffffff')
        self.result_title.pack(pady=20)
        
        self.table_frame = tk.Frame(self.results_frame, bg='#ffffff')
        self.table_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Kolom khusus untuk belanja persediaan
        columns = ('Kode Rekening', 'Kode Kegiatan', 'Uraian', 'Jumlah (Rp)')
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show='headings', height=15)
        
        # Style untuk memperbesar font tabel
        style = ttk.Style()
        style.configure("Treeview", font=('Arial', 12))
        style.configure("Treeview.Heading", font=('Arial', 14, 'bold'))
        
        for col in columns:
            self.tree.heading(col, text=col)
        self.tree.column('Kode Rekening', width=150, anchor='center')
        self.tree.column('Kode Kegiatan', width=120, anchor='center')
        self.tree.column('Uraian', width=350, anchor='w')
        self.tree.column('Jumlah (Rp)', width=150, anchor='w')  # Diubah dari 'e' ke 'w' untuk rata kiri
        
        scrollbar = ttk.Scrollbar(self.table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self.summary_frame = tk.Frame(self.results_frame, bg='#ecf0f1', relief='raised', bd=1)
        self.summary_frame.pack(fill='x', padx=20, pady=10)
        
        self.total_label = tk.Label(self.summary_frame, text="", font=('Arial', 14, 'bold'), bg='#ecf0f1')
        self.total_label.pack(pady=5)

    def create_ringkasan_table(self):
        """Buat tabel khusus untuk ringkasan dengan kolom yang berbeda"""
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        self.result_title = tk.Label(self.results_frame, text="Ringkasan Anggaran", 
                                    font=('Arial', 16, 'bold'), bg='#ffffff')
        self.result_title.pack(pady=20)
        
        self.table_frame = tk.Frame(self.results_frame, bg='#ffffff')
        self.table_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Kolom khusus untuk ringkasan
        columns = ('Kategori', 'Jumlah (Rp)')
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show='headings', height=15)
        
        # Style untuk memperbesar font tabel
        style = ttk.Style()
        style.configure("Treeview", font=('Arial', 12))
        style.configure("Treeview.Heading", font=('Arial', 14, 'bold'))
        
        for col in columns:
            self.tree.heading(col, text=col)
        self.tree.column('Kategori', width=400, anchor='w')
        self.tree.column('Jumlah (Rp)', width=200, anchor='w')
        
        scrollbar = ttk.Scrollbar(self.table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self.summary_frame = tk.Frame(self.results_frame, bg='#ecf0f1', relief='raised', bd=1)
        self.summary_frame.pack(fill='x', padx=20, pady=10)
        
        # Label nama sekolah
        sekolah_label = tk.Label(self.summary_frame, text=f"SEKOLAH: {self.nama_sekolah}", 
                                font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        sekolah_label.pack(anchor='w', pady=5)
    
    def is_valid_kegiatan_format(self, kode_kegiatan):
        """Cek apakah kode kegiatan sesuai format xx.xx.xx (2 digit, titik, 2 digit, titik, 2 digit, titik)"""
        if not kode_kegiatan:
            return False
        
        # Pattern untuk format xx.xx.xx.
        pattern = r'^\d{2}\.\d{2}\.\d{2}\.$'
        return bool(re.match(pattern, kode_kegiatan))
    
    def upload_excel(self):
        file_path = filedialog.askopenfilename(title="Pilih File Excel RKAS", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                self.extract_excel_data(file_path)
                self.file_label.config(text=f"File dipilih: {file_path.split('/')[-1]}", fg='#27ae60')

                # ✅ Nonaktifkan tombol setelah upload berhasil
                self.upload_btn.config(state='disabled')

                messagebox.showinfo("Berhasil", f"File Excel berhasil diproses!\nTotal Penerimaan: Rp {self.total_penerimaan:,}\nDitemukan {len(self.budget_items)} item anggaran\nDitemukan {len(self.belanja_persediaan_items)} item belanja persediaan\nDitemukan {len(self.belanja_jasa_items)} item belanja jasa\nDitemukan {len(self.belanja_pemeliharaan_items)} item belanja pemeliharaan\nDitemukan {len(self.belanja_perjalanan_items)} item belanja perjalanan\nDitemukan {len(self.peralatan_items)} item peralatan dan mesin\nDitemukan {len(self.aset_tetap_items)} item aset tetap lainnya")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal membaca file Excel: {str(e)}")

    def extract_nama_sekolah(self, sheet):
        """Ekstrak nama sekolah dari baris 7, kolom F-AF (merged)"""
        try:
            # Ekstrak teks dari kolom F sampai AF (F=6, AF=32)
            nama_sekolah = self.extract_merged_text(sheet, 7, range(6, 33))  # F=6 sampai AF=32
            if nama_sekolah and nama_sekolah.strip():
                self.nama_sekolah = nama_sekolah.strip()
                print(f"Debug: Nama sekolah ditemukan: {self.nama_sekolah}")
            else:
                self.nama_sekolah = "Nama Sekolah Tidak Ditemukan"
        except Exception as e:
            print(f"Error extracting nama sekolah {e}")
            self.nama_sekolah = "Nama Sekolah Tidak Ditemukan"

    def extract_excel_data(self, file_path):
        """Ekstrak data dari file Excel dengan struktur spesifik RKAS"""
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Reset data
        self.budget_items = []
        self.belanja_persediaan_items = []
        self.belanja_jasa_items = []
        self.belanja_pemeliharaan_items = []
        self.belanja_perjalanan_items = []
        self.peralatan_items = [] 
        self.aset_tetap_items = []
        self.total_penerimaan = 0
        
        # Ekstrak nama sekolah - TAMBAHKAN INI
        self.extract_nama_sekolah(sheet)

        # Ekstrak total penerimaan dari baris 30, kolom I-N (merged)
        self.extract_total_penerimaan(sheet)
        
        # Ekstrak data budget berdasarkan kode kegiatan
        self.extract_budget_data(sheet)
        
        # Ekstrak data belanja persediaan berdasarkan kode rekening
        self.extract_belanja_persediaan_data(sheet)

        # Ekstrak data belanja jasa berdasarkan kode rekening
        self.extract_belanja_jasa_data(sheet)

        # Ekstrak data belanja pemeliharaan berdasarkan kode rekening
        self.extract_belanja_pemeliharaan_data(sheet)

        # Ekstrak data belanja perjalanan dinas berdasarkan kode rekening
        self.extract_belanja_perjalanan_data(sheet)

        # Ekstrak data peralatan berdasarkan kode rekening
        self.extract_peralatan_data(sheet)

        # Ekstrak data aset tetap lainnya berdasarkan kode rekening
        self.extract_aset_tetap_data(sheet)
        
        print(f"Debug: Total Penerimaan: Rp {self.total_penerimaan:,}")
        print(f"Debug: Found {len(self.belanja_persediaan_items)} belanja persediaan items")
        print(f"Debug: Found {len(self.belanja_jasa_items)} belanja jasa items")  # Tambahkan ini
        print(f"Debug: Found {len(self.belanja_pemeliharaan_items)} belanja pemeliharaan items")
        print(f"Debug: Found {len(self.belanja_perjalanan_items)} belanja perjalanan items")
        print(f"Debug: Found {len(self.peralatan_items)} peralatan items")
        print(f"Debug: Found {len(self.aset_tetap_items)} aset tetap items")
    
    def extract_total_penerimaan(self, sheet):
        """Ekstrak total penerimaan dari baris 30"""
        try:
            # Cari di baris 30, kolom I sampai N
            for col_idx in range(9, 15):  # I=9, J=10, K=11, L=12, M=13, N=14
                cell_value = sheet.cell(row=30, column=col_idx).value
                if isinstance(cell_value, (int, float)) and cell_value > 100000000:  # Minimal 100 juta
                    self.total_penerimaan = int(cell_value)
                    print(f"Debug: Total Penerimaan ditemukan di baris 30, kolom {chr(64+col_idx)}: Rp {self.total_penerimaan:,}")
                    return
            
            # Jika tidak ditemukan di baris 30, cari di sekitar baris tersebut
            for row_idx in range(28, 33):  # Baris 28-32
                for col_idx in range(9, 15):  # Kolom I-N
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(cell_value, (int, float)) and cell_value > 100000000:
                        self.total_penerimaan = int(cell_value)
                        print(f"Debug: Total Penerimaan ditemukan di baris {row_idx}, kolom {chr(64+col_idx)}: Rp {self.total_penerimaan:,}")
                        return
        except Exception as e:
            print(f"Error extracting total penerimaan: {e}")
        
        # Default jika tidak ditemukan
        if not self.total_penerimaan:
            self.total_penerimaan = 799500000  # Sesuai dengan gambar
            print("Debug: Menggunakan default total penerimaan: Rp 799,500,000")
    
    def extract_budget_data(self, sheet):
        """Ekstrak data budget berdasarkan kode kegiatan di kolom G"""
        # Semua kode yang dicari
        all_target_codes = (self.kategori_kode['buku'] + 
                           self.kategori_kode['sarana_prasarana'] + 
                           self.kategori_kode['honor'])
        
        print(f"Debug: Mencari kode: {all_target_codes}")
        
        # Iterasi semua baris untuk mencari kode kegiatan
        for row_idx in range(1, sheet.max_row + 1):
            # Baca kode kegiatan dari kolom G (index 7)
            kode_cell = sheet.cell(row=row_idx, column=7)  # Kolom G
            kode_value = str(kode_cell.value) if kode_cell.value else ""
            kode_value = kode_value.strip()
            
            if not kode_value or kode_value == "None":
                continue
            
            # Cek apakah kode ini cocok dengan yang dicari
            for target_code in all_target_codes:
                if target_code in kode_value:
                    print(f"Debug: Menemukan kode {target_code} di baris {row_idx}: {kode_value}")
                    
                    # Ekstrak uraian dari kolom H-M (merged)
                    uraian = self.extract_merged_text(sheet, row_idx, range(8, 14))  # H=8, I=9, J=10, K=11, L=12, M=13
                    
                    # Ekstrak jumlah dari kolom N-O (merged)
                    jumlah = self.extract_merged_number(sheet, row_idx, range(14, 16))  # N=14, O=15
                    
                    if uraian and jumlah > 0:
                        # Cek apakah kode sudah ada (hindari duplikasi)
                        existing = next((item for item in self.budget_items if item['kode'] == target_code), None)
                        if not existing:
                            self.budget_items.append({
                                'kode': target_code,
                                'uraian': uraian,
                                'jumlah': jumlah
                            })
                            print(f"Debug: Added - {target_code}: {uraian} - Rp {jumlah:,}")
                    break
    
    def extract_belanja_persediaan_data(self, sheet):
        """Ekstrak data belanja persediaan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.01'
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_code}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.extract_merged_text_strict(sheet, row_idx, range(4, 7))  # D=4, E=5, F=6
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung target code
            if target_code in kode_rekening:
                print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                
                # Ekstrak kode kegiatan dari kolom G
                kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                
                # Filter hanya kode kegiatan dengan format xx.xx.xx.
                if not self.is_valid_kegiatan_format(kode_kegiatan):
                    print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                    continue
                
                # Ekstrak uraian dari kolom H-M (merged)
                uraian = self.extract_merged_text(sheet, row_idx, range(8, 14))  # H=8, I=9, J=10, K=11, L=12, M=13
                
                # Ekstrak jumlah dari kolom N-O (merged)
                jumlah = self.extract_merged_number(sheet, row_idx, range(14, 16))  # N=14, O=15
                
                if uraian and jumlah > 0 and kode_kegiatan:
                    found_items.append({
                        'kode_rekening': kode_rekening,
                        'kode_kegiatan': kode_kegiatan,
                        'uraian': uraian,
                        'jumlah': jumlah,
                        'row': row_idx
                    })
                    print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama,
        # ambil yang paling atas (biasanya yang total/parent)
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        self.belanja_persediaan_items = filtered_items

    def extract_belanja_jasa_data(self, sheet):
        """Ekstrak data belanja jasa berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.02'
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_code}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.extract_merged_text_strict(sheet, row_idx, range(4, 7))  # D=4, E=5, F=6
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung target code
            if target_code in kode_rekening:
                print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                
                # Ekstrak kode kegiatan dari kolom G
                kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                
                # Filter hanya kode kegiatan dengan format xx.xx.xx.
                if not self.is_valid_kegiatan_format(kode_kegiatan):
                    print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                    continue
                
                # Ekstrak uraian dari kolom H-M (merged)
                uraian = self.extract_merged_text(sheet, row_idx, range(8, 14))  # H=8, I=9, J=10, K=11, L=12, M=13
                
                # Ekstrak jumlah dari kolom N-O (merged)
                jumlah = self.extract_merged_number(sheet, row_idx, range(14, 16))  # N=14, O=15
                
                if uraian and jumlah > 0 and kode_kegiatan:
                    found_items.append({
                        'kode_rekening': kode_rekening,
                        'kode_kegiatan': kode_kegiatan,
                        'uraian': uraian,
                        'jumlah': jumlah,
                        'row': row_idx
                    })
                    print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama,
        # ambil yang paling atas (biasanya yang total/parent)
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        self.belanja_jasa_items = filtered_items

    def extract_belanja_pemeliharaan_data(self, sheet):
        """Ekstrak data belanja pemeliharaan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.03'
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_code}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.extract_merged_text_strict(sheet, row_idx, range(4, 7))  # D=4, E=5, F=6
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung target code
            if target_code in kode_rekening:
                print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                
                # Ekstrak kode kegiatan dari kolom G
                kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                
                # Filter hanya kode kegiatan dengan format xx.xx.xx.
                if not self.is_valid_kegiatan_format(kode_kegiatan):
                    print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                    continue
                
                # Ekstrak uraian dari kolom H-M (merged)
                uraian = self.extract_merged_text(sheet, row_idx, range(8, 14))  # H=8, I=9, J=10, K=11, L=12, M=13
                
                # Ekstrak jumlah dari kolom N-O (merged)
                jumlah = self.extract_merged_number(sheet, row_idx, range(14, 16))  # N=14, O=15
                
                if uraian and jumlah > 0 and kode_kegiatan:
                    found_items.append({
                        'kode_rekening': kode_rekening,
                        'kode_kegiatan': kode_kegiatan,
                        'uraian': uraian,
                        'jumlah': jumlah,
                        'row': row_idx
                    })
                    print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama,
        # ambil yang paling atas (biasanya yang total/parent)
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        self.belanja_pemeliharaan_items = filtered_items

    def extract_belanja_perjalanan_data(self, sheet):
        """Ekstrak data belanja perjalanan dinas berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.1.02.04'
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_code}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.extract_merged_text_strict(sheet, row_idx, range(4, 7))  # D=4, E=5, F=6
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung target code
            if target_code in kode_rekening:
                print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                
                # Ekstrak kode kegiatan dari kolom G
                kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                
                # Filter hanya kode kegiatan dengan format xx.xx.xx.
                if not self.is_valid_kegiatan_format(kode_kegiatan):
                    print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                    continue
                
                # Ekstrak uraian dari kolom H-M (merged)
                uraian = self.extract_merged_text(sheet, row_idx, range(8, 14))  # H=8, I=9, J=10, K=11, L=12, M=13
                
                # Ekstrak jumlah dari kolom N-O (merged)
                jumlah = self.extract_merged_number(sheet, row_idx, range(14, 16))  # N=14, O=15
                
                if uraian and jumlah > 0 and kode_kegiatan:
                    found_items.append({
                        'kode_rekening': kode_rekening,
                        'kode_kegiatan': kode_kegiatan,
                        'uraian': uraian,
                        'jumlah': jumlah,
                        'row': row_idx
                    })
                    print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama,
        # ambil yang paling atas (biasanya yang total/parent)
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        self.belanja_perjalanan_items = filtered_items

    def extract_peralatan_data(self, sheet):
        """Ekstrak data peralatan berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_code = '5.2.02'
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_code}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.extract_merged_text_strict(sheet, row_idx, range(4, 7))  # D=4, E=5, F=6
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung target code
            if target_code in kode_rekening:
                print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                
                # Ekstrak kode kegiatan dari kolom G
                kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                
                # Filter hanya kode kegiatan dengan format xx.xx.xx.
                if not self.is_valid_kegiatan_format(kode_kegiatan):
                    print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                    continue
                
                # Ekstrak uraian dari kolom H-M (merged)
                uraian = self.extract_merged_text(sheet, row_idx, range(8, 14))  # H=8, I=9, J=10, K=11, L=12, M=13
                
                # Ekstrak jumlah dari kolom N-O (merged)
                jumlah = self.extract_merged_number(sheet, row_idx, range(14, 16))  # N=14, O=15
                
                if uraian and jumlah > 0 and kode_kegiatan:
                    found_items.append({
                        'kode_rekening': kode_rekening,
                        'kode_kegiatan': kode_kegiatan,
                        'uraian': uraian,
                        'jumlah': jumlah,
                        'row': row_idx
                    })
                    print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
        
        # Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama,
        # ambil yang paling atas (biasanya yang total/parent)
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        self.peralatan_items = filtered_items
    
    def extract_aset_tetap_data(self, sheet):
        """Ekstrak data aset tetap lainnya berdasarkan kode rekening di kolom D-F (merged) - STRICT VERSION"""
        target_codes = ['5.2.04', '5.2.05']
        found_items = []
        
        print(f"Debug: Mencari kode rekening yang mengandung: {target_codes}")
        
        # Iterasi semua baris untuk mencari kode rekening
        for row_idx in range(1, sheet.max_row + 1):
            # STRICT: Hanya baca kode rekening dari baris yang tepat, tanpa fallback ke baris lain
            kode_rekening = self.extract_merged_text_strict(sheet, row_idx, range(4, 7))  # D=4, E=5, F=6
            
            if not kode_rekening or kode_rekening == "None":
                continue
            
            # Cek apakah kode rekening mengandung salah satu target code
            for target_code in target_codes:
                if target_code in kode_rekening:
                    print(f"Debug: Menemukan kode rekening {kode_rekening} di baris {row_idx}")
                    
                    # Ekstrak kode kegiatan dari kolom G
                    kode_kegiatan = str(sheet.cell(row=row_idx, column=7).value or "").strip()
                    
                    # Filter hanya kode kegiatan dengan format xx.xx.xx.
                    if not self.is_valid_kegiatan_format(kode_kegiatan):
                        print(f"Debug: Skipping invalid format - {kode_kegiatan}")
                        continue
                    
                    # Ekstrak uraian dari kolom H-M (merged)
                    uraian = self.extract_merged_text(sheet, row_idx, range(8, 14))  # H=8, I=9, J=10, K=11, L=12, M=13
                    
                    # Ekstrak jumlah dari kolom N-O (merged)
                    jumlah = self.extract_merged_number(sheet, row_idx, range(14, 16))  # N=14, O=15
                    
                    if uraian and jumlah > 0 and kode_kegiatan:
                        found_items.append({
                            'kode_rekening': kode_rekening,
                            'kode_kegiatan': kode_kegiatan,
                            'uraian': uraian,
                            'jumlah': jumlah,
                            'row': row_idx
                        })
                        print(f"Debug: Found valid item - {kode_rekening} | {kode_kegiatan} | {uraian} - Rp {jumlah:,}")
                    break
        
        # Filter items: untuk kode rekening yang sama dengan kode kegiatan yang sama,
        # ambil yang paling atas (biasanya yang total/parent)
        filtered_items = []
        processed_combinations = set()
        
        # Group by kode_rekening + kode_kegiatan
        for item in sorted(found_items, key=lambda x: x['row']):
            combination_key = f"{item['kode_rekening']}_{item['kode_kegiatan']}"
            
            if combination_key not in processed_combinations:
                filtered_items.append(item)
                processed_combinations.add(combination_key)
                print(f"Debug: Added to final list - {item['kode_rekening']} | {item['kode_kegiatan']}")
        
        self.aset_tetap_items = filtered_items

    def extract_merged_text_strict(self, sheet, row_idx, col_range):
        """Ekstrak teks dari kolom yang di-merge - VERSI STRICT tanpa fallback ke baris lain"""
        text_parts = []
        
        # Hanya baca dari baris yang tepat saja
        for col_idx in col_range:
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value and str(cell_value).strip() and str(cell_value).strip() != "None":
                text_parts.append(str(cell_value).strip())
        
        # Gabungkan teks yang ditemukan
        combined_text = " ".join(text_parts).strip()
        
        return combined_text
    
    def extract_merged_text(self, sheet, row_idx, col_range):
        """Ekstrak teks dari kolom yang di-merge"""
        text_parts = []
        
        for col_idx in col_range:
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value and str(cell_value).strip() and str(cell_value).strip() != "None":
                text_parts.append(str(cell_value).strip())
        
        # Gabungkan teks yang ditemukan
        combined_text = " ".join(text_parts).strip()
        
        # Jika tidak ditemukan teks di baris yang sama, coba baris sebelum/sesudah (untuk handling merge cell)
        if not combined_text:
            for offset in [-1, 1, -2, 2]:
                if row_idx + offset > 0:
                    for col_idx in col_range:
                        cell_value = sheet.cell(row=row_idx + offset, column=col_idx).value
                        if cell_value and str(cell_value).strip() and str(cell_value).strip() != "None":
                            combined_text = str(cell_value).strip()
                            break
                    if combined_text:
                        break
        
        return combined_text
    
    def extract_merged_number(self, sheet, row_idx, col_range):
        """Ekstrak angka dari kolom yang di-merge"""
        for col_idx in col_range:
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_value, (int, float)) and cell_value > 0:
                return int(cell_value)
        
        # Jika tidak ditemukan di baris yang sama, coba baris sebelum/sesudah
        for offset in [-1, 1, -2, 2]:
            if row_idx + offset > 0:
                for col_idx in col_range:
                    cell_value = sheet.cell(row=row_idx + offset, column=col_idx).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        return int(cell_value)
        
        return 0
    
    def filter_budget_by_codes(self, codes: List[str]) -> List[Dict]:
        """Filter budget berdasarkan kode kategori"""
        filtered_items = []
        for item in self.budget_items:
            for code in codes:
                if item['kode'].startswith(code[:5]):  # Match first 5 characters
                    filtered_items.append(item)
                    break
        
        return sorted(filtered_items, key=lambda x: x['kode'])
    
    def display_results(self, title: str, items: List[Dict], category_name: str):
        """Tampilkan hasil ke tabel"""
        self.result_title.config(text=title)
        
        # Hapus data lama
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        total_jumlah = 0
        if items:
            for item in items:
                formatted_jumlah = f"Rp {item['jumlah']:,}".replace(',', '.')
                self.tree.insert('', 'end', values=(item['kode'], item['uraian'], formatted_jumlah))
                total_jumlah += item['jumlah']
        else:
            self.tree.insert('', 'end', values=('', 'Tidak ada data ditemukan untuk kategori ini', 'Rp 0'))
        
        # Hitung persentase
        percentage = (total_jumlah / self.total_penerimaan) * 100 if self.total_penerimaan else 0
        
        # Update summary
        # UBAH BAGIAN INI - Layout horizontal
        # Frame kiri untuk nama sekolah
        left_frame = tk.Frame(self.summary_frame, bg='#ecf0f1')
        left_frame.pack(side='left', fill='y', padx=(10, 20))
        
        sekolah_label = tk.Label(left_frame, text=f"SEKOLAH {self.nama_sekolah}", 
                                font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        sekolah_label.pack(anchor='w')
        self.total_label.config(text=f"Total Alokasi {category_name}: Rp {total_jumlah:,}".replace(',', '.'))
        self.percentage_label.config(text=f"Persentase dari Total Penerimaan: {percentage:.2f}%")
    
    def display_belanja_persediaan_results(self, items: List[Dict]):
        """Tampilkan hasil belanja persediaan ke tabel khusus"""
        # Buat tabel khusus
        self.create_belanja_persediaan_table()
        
        total_jumlah = 0
        if items:
            for item in items:
                formatted_jumlah = f"Rp {item['jumlah']:,}".replace(',', '.')
                self.tree.insert('', 'end', values=(
                    item['kode_rekening'], 
                    item['kode_kegiatan'], 
                    item['uraian'], 
                    formatted_jumlah
                ))
                total_jumlah += item['jumlah']
        else:
            self.tree.insert('', 'end', values=('', '', 'Tidak ada data ditemukan untuk kategori ini', 'Rp 0'))
        
        # Update summary (tanpa percentage)
        # UBAH BAGIAN INI - Layout horizontal
        # Frame kiri untuk nama sekolah
        left_frame = tk.Frame(self.summary_frame, bg='#ecf0f1')
        left_frame.pack(side='left', fill='y', padx=(10, 20))
        
        sekolah_label = tk.Label(left_frame, text=f"SEKOLAH {self.nama_sekolah}", 
                                font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        sekolah_label.pack(anchor='w')
        self.total_label.config(text=f"Total Pakai Habis: Rp {total_jumlah:,}".replace(',', '.'))

    def display_belanja_jasa_results(self, items: List[Dict]):
        """Tampilkan hasil belanja jasa ke tabel khusus dengan summary yang diperluas"""
        # Buat tabel khusus
        self.create_belanja_persediaan_table()
        self.result_title.config(text="Rincian Belanja Jasa (5.1.02.02)")
        
        total_belanja_jasa = 0
        if items:
            for item in items:
                formatted_jumlah = f"Rp {item['jumlah']:,}".replace(',', '.')
                self.tree.insert('', 'end', values=(
                    item['kode_rekening'], 
                    item['kode_kegiatan'], 
                    item['uraian'], 
                    formatted_jumlah
                ))
                total_belanja_jasa += item['jumlah']
        else:
            self.tree.insert('', 'end', values=('', '', 'Tidak ada data ditemukan untuk kategori ini', 'Rp 0'))
        
        # Hitung total honor dari kategori honor
        honor_items = self.filter_budget_by_codes(self.kategori_kode['honor'])
        total_honor = sum(item['jumlah'] for item in honor_items)
        
        # Hitung jasa sesungguhnya
        jasa_sesungguhnya = total_belanja_jasa - total_honor
        
        # Update summary dengan informasi lengkap - urutan vertikal
        # Baris 1: Total Belanja Jasa
        self.total_label.config(text=f"Total Belanja Jasa: Rp {total_belanja_jasa:,}".replace(',', '.'))
        
        # Baris 2: Pembayaran Honor
        if hasattr(self, 'honor_label'):
            self.honor_label.destroy()
        self.honor_label = tk.Label(self.summary_frame, 
                                text=f"Pembayaran Honor: Rp {total_honor:,}".replace(',', '.'), 
                                font=('Arial', 14, 'bold'), bg='#ecf0f1')
        self.honor_label.pack(pady=5)
        
        # Baris 3: Jasa Sesungguhnya
        if hasattr(self, 'jasa_sesungguhnya_label'):
            self.jasa_sesungguhnya_label.destroy()
        self.jasa_sesungguhnya_label = tk.Label(self.summary_frame, 
                                            text=f"Jasa Sesungguhnya: Rp {jasa_sesungguhnya:,}".replace(',', '.'), 
                                            font=('Arial', 14, 'bold'), bg='#ecf0f1', fg='#e74c3c')
        self.jasa_sesungguhnya_label.pack(pady=5)
        
        # Baris 4: Label Sekolah (paling kiri)
        if hasattr(self, 'sekolah_label'):
            self.sekolah_label.destroy()
        self.sekolah_label = tk.Label(self.summary_frame, 
                                    text=f"SEKOLAH: {self.nama_sekolah}", 
                                    font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        self.sekolah_label.pack(anchor='w', pady=(10, 5))  # anchor='w' untuk posisi kiri, padding atas untuk jarak
    
    def display_belanja_pemeliharaan_results(self, items: List[Dict]):
        """Tampilkan hasil belanja pemeliharaan ke tabel khusus"""
        # Buat tabel khusus
        self.create_belanja_persediaan_table()
        
        # Update title
        self.result_title.config(text="Rincian Belanja Pemeliharaan (5.1.02.03)")
        
        total_jumlah = 0
        if items:
            for item in items:
                formatted_jumlah = f"Rp {item['jumlah']:,}".replace(',', '.')
                self.tree.insert('', 'end', values=(
                    item['kode_rekening'], 
                    item['kode_kegiatan'], 
                    item['uraian'], 
                    formatted_jumlah
                ))
                total_jumlah += item['jumlah']
        else:
            self.tree.insert('', 'end', values=('', '', 'Tidak ada data ditemukan untuk kategori ini', 'Rp 0'))
        
        # Update summary (tanpa percentage)
        # UBAH BAGIAN INI - Layout horizontal
        # Frame kiri untuk nama sekolah
        left_frame = tk.Frame(self.summary_frame, bg='#ecf0f1')
        left_frame.pack(side='left', fill='y', padx=(10, 20))
        
        sekolah_label = tk.Label(left_frame, text=f"SEKOLAH {self.nama_sekolah}", 
                                font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        sekolah_label.pack(anchor='w')
        self.total_label.config(text=f"Total Belanja Pemeliharaan: Rp {total_jumlah:,}".replace(',', '.'))

    def display_belanja_perjalanan_results(self, items: List[Dict]):
        """Tampilkan hasil belanja perjalanan dinas ke tabel khusus"""
        # Buat tabel khusus
        self.create_belanja_persediaan_table()
        self.result_title.config(text="Rincian Perjalanan Dinas (5.1.02.04)")
        
        total_jumlah = 0
        if items:
            for item in items:
                formatted_jumlah = f"Rp {item['jumlah']:,}".replace(',', '.')
                self.tree.insert('', 'end', values=(
                    item['kode_rekening'], 
                    item['kode_kegiatan'], 
                    item['uraian'], 
                    formatted_jumlah
                ))
                total_jumlah += item['jumlah']
        else:
            self.tree.insert('', 'end', values=('', '', 'Tidak ada data ditemukan untuk kategori ini', 'Rp 0'))
        
        # Update summary (tanpa percentage)
        # UBAH BAGIAN INI - Layout horizontal
        # Frame kiri untuk nama sekolah
        left_frame = tk.Frame(self.summary_frame, bg='#ecf0f1')
        left_frame.pack(side='left', fill='y', padx=(10, 20))
        
        sekolah_label = tk.Label(left_frame, text=f"SEKOLAH {self.nama_sekolah}", 
                                font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        sekolah_label.pack(anchor='w')
        self.total_label.config(text=f"Total Belanja Perjalanan Dinas: Rp {total_jumlah:,}".replace(',', '.'))

    def display_peralatan_results(self, items: List[Dict]):
        """Tampilkan hasil peralatan ke tabel khusus"""
        # Buat tabel khusus
        self.create_belanja_persediaan_table()
        self.result_title.config(text="Rincian Peralatan dan Mesin (5.2.02)")
        
        total_jumlah = 0
        if items:
            for item in items:
                formatted_jumlah = f"Rp {item['jumlah']:,}".replace(',', '.')
                self.tree.insert('', 'end', values=(
                    item['kode_rekening'], 
                    item['kode_kegiatan'], 
                    item['uraian'], 
                    formatted_jumlah
                ))
                total_jumlah += item['jumlah']
        else:
            self.tree.insert('', 'end', values=('', '', 'Tidak ada data ditemukan untuk kategori ini', 'Rp 0'))
        
        # Update summary (tanpa percentage)
        # UBAH BAGIAN INI - Layout horizontal
        # Frame kiri untuk nama sekolah
        left_frame = tk.Frame(self.summary_frame, bg='#ecf0f1')
        left_frame.pack(side='left', fill='y', padx=(10, 20))
        
        sekolah_label = tk.Label(left_frame, text=f"SEKOLAH {self.nama_sekolah}", 
                                font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        sekolah_label.pack(anchor='w')
        self.total_label.config(text=f"Total Peralatan: Rp {total_jumlah:,}".replace(',', '.'))

    def display_aset_tetap_results(self, items: List[Dict]):
        """Tampilkan hasil aset tetap lainnya ke tabel khusus"""
        # Buat tabel khusus
        self.create_belanja_persediaan_table()
        
        # Update title
        self.result_title.config(text="Rincian Aset Tetap Lainnya (5.2.04 & 5.2.05)")
        
        total_jumlah = 0
        if items:
            for item in items:
                formatted_jumlah = f"Rp {item['jumlah']:,}".replace(',', '.')
                self.tree.insert('', 'end', values=(
                    item['kode_rekening'], 
                    item['kode_kegiatan'], 
                    item['uraian'], 
                    formatted_jumlah
                ))
                total_jumlah += item['jumlah']
        else:
            self.tree.insert('', 'end', values=('', '', 'Tidak ada data ditemukan untuk kategori ini', 'Rp 0'))
        
        # Update summary (tanpa percentage)
        # UBAH BAGIAN INI - Layout horizontal
        # Frame kiri untuk nama sekolah
        left_frame = tk.Frame(self.summary_frame, bg='#ecf0f1')
        left_frame.pack(side='left', fill='y', padx=(10, 20))
        
        sekolah_label = tk.Label(left_frame, text=f"SEKOLAH {self.nama_sekolah}", 
                                font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50')
        sekolah_label.pack(anchor='w')
        self.total_label.config(text=f"Total Aset Tetap Lainnya: Rp {total_jumlah:,}".replace(',', '.'))

    def show_buku(self):
        if not self.budget_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.create_results_table()  # Reset ke tabel normal
        items = self.filter_budget_by_codes(self.kategori_kode['buku'])
        self.display_results("Rincian Anggaran Buku", items, "Buku")
    
    def show_sarana_prasarana(self):
        if not self.budget_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.create_results_table()  # Reset ke tabel normal
        items = self.filter_budget_by_codes(self.kategori_kode['sarana_prasarana'])
        self.display_results("Rincian Anggaran Sarana & Prasarana", items, "Sarana & Prasarana")
    
    def show_honor(self):
        if not self.budget_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.create_results_table()  # Reset ke tabel normal
        items = self.filter_budget_by_codes(self.kategori_kode['honor'])
        self.display_results("Rincian Anggaran Honor", items, "Honor")
    
    def show_belanja_persediaan(self):
        if not self.belanja_persediaan_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.display_belanja_persediaan_results(self.belanja_persediaan_items)

    def show_belanja_jasa(self):
        if not self.belanja_jasa_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.display_belanja_jasa_results(self.belanja_jasa_items)

    def show_belanja_pemeliharaan(self):
        if not self.belanja_pemeliharaan_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.display_belanja_pemeliharaan_results(self.belanja_pemeliharaan_items)

    def show_belanja_perjalanan(self):
        if not self.belanja_perjalanan_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.display_belanja_perjalanan_results(self.belanja_perjalanan_items)

    def show_peralatan(self):
        if not self.peralatan_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.display_peralatan_results(self.peralatan_items)

    def show_aset_tetap(self):
        if not self.aset_tetap_items:
            messagebox.showwarning("Peringatan", "Data dalam kategori tersebut tidak ada atau file belum diupload!")
            return
        self.display_aset_tetap_results(self.aset_tetap_items)

    def show_ringkasan(self):
        if not self.excel_data and self.total_penerimaan == 0:
            messagebox.showwarning("Peringatan", "File belum diupload!")
            return
        
        # Buat tabel khusus ringkasan
        self.create_ringkasan_table()
        
        # Hitung semua nilai yang diperlukan
        total_belanja_persediaan = sum(item['jumlah'] for item in self.belanja_persediaan_items)
        
        # Honor dari kategori budget
        honor_items = self.filter_budget_by_codes(self.kategori_kode['honor'])
        total_honor = sum(item['jumlah'] for item in honor_items)
        
        # Belanja jasa dan jasa sesungguhnya
        total_belanja_jasa = sum(item['jumlah'] for item in self.belanja_jasa_items)
        jasa_sesungguhnya = total_belanja_jasa - total_honor
        
        total_pemeliharaan = sum(item['jumlah'] for item in self.belanja_pemeliharaan_items)
        total_perjalanan = sum(item['jumlah'] for item in self.belanja_perjalanan_items)
        total_peralatan = sum(item['jumlah'] for item in self.peralatan_items)
        total_aset_tetap = sum(item['jumlah'] for item in self.aset_tetap_items)
        
        # Hitung belanja persediaan yang dimaksud dalam ringkasan
        belanja_persediaan_ringkasan = (self.total_penerimaan - total_honor - 
                                    jasa_sesungguhnya - total_pemeliharaan - total_perjalanan)
        
        belanja_modal = total_peralatan + total_aset_tetap
        total_anggaran = total_belanja_persediaan + total_peralatan + total_aset_tetap
        
        # Data ringkasan dengan background color info
        ringkasan_data = [
            ("PAGU TAHUN 2025", self.total_penerimaan, True),  # True = background hijau
            ("BELANJA OPERASI", total_belanja_persediaan, True),
            ("  BELANJA HONOR", total_honor, False),
            ("  BELANJA JASA", jasa_sesungguhnya, False),
            ("  BELANJA PEMELIHARAAN", total_pemeliharaan, False),
            ("  BELANJA PERJALANAN", total_perjalanan, False),
            ("  BELANJA PERSEDIAAN", belanja_persediaan_ringkasan, True),
            ("BELANJA MODAL", belanja_modal, True),
            ("  PERALATAN DAN MESIN", total_peralatan, False),
            ("  ASET TETAP LAINNYA", total_aset_tetap, False),
            ("TOTAL ANGGARAN", total_anggaran, True)
        ]
        
        # Masukkan data ke tabel
        for kategori, jumlah, is_highlight in ringkasan_data:
            formatted_jumlah = f"Rp {jumlah:,}".replace(',', '.')
            item_id = self.tree.insert('', 'end', values=(kategori, formatted_jumlah))
            
            # Set background color untuk item yang di-highlight
            if is_highlight:
                self.tree.set(item_id, 'Kategori', kategori)
                self.tree.set(item_id, 'Jumlah (Rp)', formatted_jumlah)
                # Konfigurasi tag untuk background hijau
                self.tree.tag_configure('highlight', background='#2ecc71', foreground='white')
                self.tree.item(item_id, tags=('highlight',))
    
    def reset_data(self):
        self.excel_data = None
        self.total_penerimaan = 0
        self.budget_items = []
        self.belanja_persediaan_items = []
        self.belanja_jasa_items = []
        self.belanja_pemeliharaan_items = []
        self.belanja_perjalanan_items = []
        self.peralatan_items = []
        self.aset_tetap_items = []

        # Bersihkan tampilan hasil
        for widget in self.results_frame.winfo_children():
            widget.destroy()

        # Kosongkan label file
        self.file_label.config(text="Belum ada file yang dipilih")

        # ✅ Aktifkan kembali tombol upload
        self.upload_btn.config(state='normal')

        messagebox.showinfo("Reset", "Data berhasil dibersihkan.")



def main():
    root = tk.Tk()
    app = BOSBudgetAnalyzer(root)
    root.mainloop()

if __name__ == "__main__":
    main()